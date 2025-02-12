import requests
from info import info
import pdb
import json
import time
import pandas as pd
from Summoner import Summoner
from openpyxl import load_workbook



#def toXlsx(cObjects, patchNum, summonerTag):



if __name__ == "__main__":
    lnterprice = Summoner("Interprice", "NA1")
    lnterprice.getAllMatches("14.23")
    lnterprice.parseChampionPool()
    championObj = lnterprice.getCObjects()
    lnterprice.getTag()


    championList = {}
    for cObj in championObj:
        cObjDict = cObj.getDict()
        championName, id = list(cObjDict.keys())[0]
        info = list(cObjDict.values())[0]
        for stat in info.keys():
            if type(info[stat]) == float:
                info[stat] = round(info[stat], 0)
                #pdb.set_trace()
            if type(info[stat]) == tuple:
                info[stat] = (round(info[stat][0], 0), round(info[stat][1], 0), round(info[stat][2], 0))
        championList[championName] = info
    pdb.set_trace()
    df = pd.DataFrame.from_dict(championList, orient="index")
    df['winRate'] = df['winRate'] * 100
    df['KP'] = (round(df['KP'] * 100, 0)).astype(str) + "%"
    df.to_excel("excel/output.xlsx", engine="openpyxl")
    wb = load_workbook("excel/output.xlsx")
    ws = wb.active
    ws["A1"] = "Sinvu#NA1"
    wb.save("excel/output.xlsx")